#!/usr/bin/env python3
"""
Fetch 7709 (CSOP SK Hynix Daily (2x) Leveraged Product, HKEX:07709) daily
Asset Under Management (AUM), NAV, units outstanding from HKEXnews
'Trading Information of Leveraged and Inverse Products' XLSX disclosures.

Data source: HKEXnews title-search public API (no auth, returns JSON).
Each trading day CSOP files one XLSX covering all its L&I products; we
locate the 7709 column and read NAV / units / AUM / premium.

Usage:
  python fetch_7709_aum.py                 # incremental: only missing dates
  python fetch_7709_aum.py --force         # re-download & re-parse everything
  python fetch_7709_aum.py --test          # dry-run on 3 sampled dates
  python fetch_7709_aum.py --history PATH  # custom output json path
"""
import json
import os
import re
import sys
import argparse
import urllib.request
from openpyxl import load_workbook

STOCK_ID = "1000276797"
API_URL = "https://www1.hkexnews.hk/search/titleSearchServlet.do"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_HISTORY = os.path.join(HERE, "..", "data", "7709_aum_history.json")
CACHE_DIR = os.path.join(HERE, "..", "data", "cache_7709")


def fetch_xlsx_links():
    """Call HKEXnews title search, return list of {date, release, link}."""
    params = (f"lang=EN&stockId={STOCK_ID}&category=0&market=SEHK"
              f"&searchType=0&rowRange=300")
    url = f"{API_URL}?{params}"
    req = urllib.request.Request(
        url, headers={"User-Agent": UA, "Accept": "application/json, */*"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.loads(r.read().decode("utf-8"))
    rows = json.loads(data["result"])
    out = []
    for r in rows:
        if r.get("FILE_TYPE") != "XLSX":
            continue
        txt = (r.get("SHORT_TEXT") or "") + (r.get("LONG_TEXT") or "")
        if "Trading Information" not in txt:
            continue
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", r.get("DATE_TIME", ""))
        if not m:
            continue
        dd, mm, yy = m.groups()
        date = f"{yy}-{mm}-{dd}"
        link = "https://www1.hkexnews.hk" + r["FILE_LINK"]
        out.append({"date": date, "release": r["DATE_TIME"], "link": link})
    return out


def download(link, cache_path):
    if os.path.exists(cache_path) and os.path.getsize(cache_path) > 0:
        return cache_path
    req = urllib.request.Request(
        link, headers={"User-Agent": UA, "Referer": "https://www1.hkexnews.hk/"})
    with urllib.request.urlopen(req, timeout=30) as r:
        content = r.read()
    with open(cache_path, "wb") as f:
        f.write(content)
    return cache_path


def parse_xlsx(path):
    """Locate 7709 column in the horizontal L&I table and extract metrics."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]

    # Stock Code row -> column index of "7709"
    code_col = None
    for row in grid:
        for ci, v in enumerate(row):
            if str(v).strip() == "7709":
                code_col = ci
                break
        if code_col is not None:
            break
    if code_col is None:
        return None
    # In this horizontal table each product's code, value and the left
    # neighbour (currency) all line up in the same / adjacent columns:
    #   <currency> | <code == value>
    val_col = code_col          # 7709 code & value share the same column
    ccy_col = code_col - 1      # currency sits one column to the left

    def get(label_kw, require_sub=None):
        for row in grid:
            label = str(row[0]).strip() if row and row[0] is not None else ""
            if label_kw not in label:
                continue
            if require_sub and require_sub not in label:
                continue
            val = row[val_col] if val_col < len(row) else None
            ccy = row[ccy_col] if 0 <= ccy_col < len(row) else None
            return val, ccy
        return None, None

    nav, nav_ccy = get("N.A.V. per Unit in Trading Currency")
    units, _ = get("Total Units Outstanding", "Hong Kong")
    aum, aum_ccy = get("Asset Under Management", "Hong Kong")
    premium, _ = get("Premium / Discount")

    def num(x):
        try:
            return float(str(x).replace(",", ""))
        except (TypeError, ValueError):
            return None

    nav_n = num(nav)
    prem_n = num(premium)
    # 溢价率定义 = (收市价 - NAV) / NAV，故收市价 = NAV * (1 + 溢价率/100)
    # 反推的收市价与港交所披露口径一致，用于计算每日涨跌幅（同源、官方）
    close = None
    if nav_n is not None and prem_n is not None:
        close = round(nav_n * (1 + prem_n / 100), 4)

    return {
        "nav": nav_n,
        "nav_ccy": (str(nav_ccy).strip() if nav_ccy else None),
        "units": num(units),
        "aum_usd": num(aum),
        "aum_ccy": (str(aum_ccy).strip() if aum_ccy else None),
        "premium": prem_n,
        "close_price": close,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--history", default=DEFAULT_HISTORY)
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--test", action="store_true")
    args = ap.parse_args()

    os.makedirs(CACHE_DIR, exist_ok=True)

    links = fetch_xlsx_links()
    # dedupe by date, keep latest release
    by_date = {}
    for l in links:
        if l["date"] not in by_date or l["release"] > by_date[l["date"]]["release"]:
            by_date[l["date"]] = l

    hist = {}
    if os.path.exists(args.history) and not args.force:
        with open(args.history) as f:
            for rec in json.load(f):
                hist[rec["date"]] = rec

    to_fetch = [l for d, l in by_date.items() if d not in hist or args.force]
    if args.test:
        n = len(to_fetch)
        idx = sorted(set([0, n // 2, n - 1]))
        to_fetch = [to_fetch[i] for i in idx if 0 <= i < n]

    print(f"links={len(by_date)} existing={len(hist)} to_fetch={len(to_fetch)}")
    for l in to_fetch:
        cache = os.path.join(CACHE_DIR, l["date"] + ".xlsx")
        try:
            download(l["link"], cache)
            parsed = parse_xlsx(cache)
            if parsed:
                hist[l["date"]] = {
                    "date": l["date"], "release": l["release"],
                    "link": l["link"], **parsed}
                print(f"  OK  {l['date']}  aum={parsed['aum_usd']}  "
                      f"nav={parsed['nav']}{parsed['nav_ccy']}  "
                      f"units={parsed['units']}")
            else:
                print(f"  SKIP {l['date']} (7709 column not found)")
        except Exception as e:
            print(f"  ERR {l['date']}: {e}")

    out = [hist[d] for d in sorted(hist.keys())]
    # 按日期升序计算每日涨跌幅（基于反推收市价序列）
    prev_close = None
    for rec in out:
        cp = rec.get("close_price")
        if prev_close is not None and cp is not None and prev_close > 0:
            rec["daily_change_pct"] = round((cp - prev_close) / prev_close * 100, 2)
        else:
            rec["daily_change_pct"] = None
        if cp is not None:
            prev_close = cp
    os.makedirs(os.path.dirname(os.path.abspath(args.history)), exist_ok=True)
    with open(args.history, "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"wrote {len(out)} records -> {args.history}")


if __name__ == "__main__":
    main()
